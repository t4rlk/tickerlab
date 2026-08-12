# -*- coding: utf-8 -*-
"""
Export academique PDF + Excel dans un dossier persistant (exports/).

Tables produites (ordre canonique Box-Jenkins) :
  1. Statistiques descriptives — prix log et rendements
  2. Correlogrammes — prix log, rendements (figure + table 30 lags)
  3. Tests de racine unitaire — ADF / PP / KPSS (3 specs + synthese)
  4. Modelisation ARIMA — grille AIC p*q, estimation, residus
  5. Test ARCH-LM sur residus ARIMA
  6. Modelisation GARCH — grille comparative, estimation, residus standardises
  7. VaR & TVaR
  8. Backtesting (Kupiec / Christoffersen)

Le dossier exports/ n'est JAMAIS efface par force_clean.
Flag : config.export_academique.enabled  (defaut : false)
"""
from __future__ import annotations

import logging
import warnings
import math
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

_log = logging.getLogger('tickerlab.export_academique')

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    BaseDocTemplate, Frame, Image, PageBreak, PageTemplate,
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from tickerlab.core.rapport._helpers import (
    _tableau_eviews, _correlogramme_eviews,
    _h1, _h2, _p, _caption, _spacer, _note, _fmt, _fmt_pval, _fmt_signif,
    LARG_U, PAGE_W, PAGE_H, MARGE_G, MARGE_D, MARGE_H, MARGE_B, nom_actif,
)
from tickerlab.core.rapport._stats import (
    stats_desc, adf_complet, pp_complet, kpss_complet, arch_lm,
)
from tickerlab.core.rapport._sections import (
    _tab_adf, _tab_pp, _tab_kpss, _tab_synthese, _tab_stats_desc,
    _dec_adf_pp, _dec_kpss,
)

warnings.filterwarnings('ignore')


# ── Constantes mise en page ───────────────────────────────────────────────────

_BLEU_FONCE  = colors.HexColor('#1F3864')
_BLEU_CLAIR  = colors.HexColor('#BDD7EE')
_GRIS_LIGNE  = colors.HexColor('#F2F2F2')
_ROUGE_SIG   = colors.HexColor('#C00000')


# ── Styles locaux ─────────────────────────────────────────────────────────────

def _styles() -> dict:
    base = getSampleStyleSheet()
    def _s(n, **kw):
        return ParagraphStyle(n, parent=base['Normal'], **kw)
    return {
        'H1':      _s('EA_H1', fontSize=13, fontName='Helvetica-Bold',
                       spaceBefore=14, spaceAfter=6, textColor=_BLEU_FONCE),
        'H2':      _s('EA_H2', fontSize=11, fontName='Helvetica-Bold',
                       spaceBefore=9,  spaceAfter=4, textColor=_BLEU_FONCE),
        'Body':    _s('EA_Body', fontSize=10, fontName='Helvetica',
                       spaceAfter=4, leading=14),
        'Caption': _s('EA_Cap', fontSize=8, fontName='Helvetica-Oblique',
                       alignment=TA_CENTER, spaceAfter=6,
                       textColor=colors.HexColor('#595959')),
        'Note':    _s('EA_Note', fontSize=8, fontName='Helvetica',
                       spaceAfter=8,  textColor=colors.HexColor('#595959')),
        'Titre':   _s('EA_Titre', fontSize=20, fontName='Helvetica-Bold',
                       alignment=TA_CENTER, spaceAfter=10, textColor=_BLEU_FONCE),
        'SousTitre': _s('EA_ST', fontSize=13, fontName='Helvetica',
                         alignment=TA_CENTER, spaceAfter=6,
                         textColor=colors.HexColor('#595959')),
    }


_ST = _styles()


def _h(level: int, text: str) -> Paragraph:
    return Paragraph(text, _ST['H1'] if level == 1 else _ST['H2'])


def _pb() -> PageBreak:
    return PageBreak()


def _sp(h: float = 0.4) -> Spacer:
    return Spacer(1, h * cm)


def _cap(text: str) -> Paragraph:
    return Paragraph(f'<i>{text}</i>', _ST['Caption'])


def _nt(text: str) -> Paragraph:
    return Paragraph(text, _ST['Note'])


def _embed(fig, w_cm: float = 15.5, h_cm: float = 4.5) -> Image:
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=200, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return Image(buf, width=w_cm * cm, height=h_cm * cm)


# ── Tableau EViews academique ─────────────────────────────────────────────────

def _tab(titre: str, cols: list, rows: list, note: str = '',
         col_widths=None, extra=None, col0_label: bool = True) -> list:
    """Wrapper autour de _tableau_eviews avec la theme academique bleu."""
    n_cols = len(cols)

    if col_widths is None:
        if n_cols == 2:
            col_widths = [LARG_U * 0.55, LARG_U * 0.45]
        elif n_cols == 3:
            col_widths = [LARG_U * 0.40] + [LARG_U * 0.30] * 2
        else:
            first_w = LARG_U * 0.28
            other_w = (LARG_U - first_w) / (n_cols - 1)
            col_widths = [first_w] + [other_w] * (n_cols - 1)

    th = getSampleStyleSheet()
    th_style = ParagraphStyle('_TH', parent=th['Normal'], fontSize=9,
                               fontName='Helvetica-Bold', alignment=TA_CENTER,
                               textColor=colors.white)

    data = [[Paragraph(f'<b>{titre}</b>', th_style)] + [''] * (n_cols - 1),
            [Paragraph(c, ParagraphStyle('_HD', parent=th['Normal'], fontSize=9,
                                          fontName='Helvetica-Bold', alignment=TA_CENTER))
             for c in cols]]
    data.extend(rows)
    n_rows = len(rows)

    cmds = [
        ('SPAN',          (0, 0), (-1, 0)),
        ('BACKGROUND',    (0, 0), (-1, 0),  _BLEU_FONCE),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
        ('VALIGN',        (0, 0), (-1, 0),  'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, 0),  6),
        ('BOTTOMPADDING', (0, 0), (-1, 0),  6),
        ('BACKGROUND',    (0, 1), (-1, 1),  _BLEU_CLAIR),
        ('ALIGN',         (0, 1), (-1, 1),  'CENTER'),
        ('FONTNAME',      (0, 1), (-1, 1),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 1), (-1, 1),  9),
        ('TEXTCOLOR',     (0, 1), (-1, 1),  colors.black),
        ('LINEBELOW',     (0, 1), (-1, 1),  1.5, colors.black),
        ('FONTSIZE',      (0, 2), (-1, -1), 9),
        ('TOPPADDING',    (0, 2), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 2), (-1, -1), 2),
        ('LINEABOVE',     (0, 0), (-1, 0),  1.0, _BLEU_FONCE),
        ('LINEBELOW',     (0, -1), (-1, -1), 1.0, _BLEU_FONCE),
        ('LINEBEFORE',    (0, 0), (0, -1),  0.5, _BLEU_FONCE),
        ('LINEAFTER',     (-1, 0), (-1, -1), 0.5, _BLEU_FONCE),
        ('INNERGRID',     (0, 1), (-1, -1), 0.25, colors.HexColor('#C9C9C9')),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
    ]

    if col0_label:
        cmds += [('FONTNAME', (0, 2), (0, -1),  'Helvetica'),
                 ('ALIGN',    (0, 2), (0, -1),  'LEFT'),
                 ('FONTNAME', (1, 2), (-1, -1), 'Courier'),
                 ('ALIGN',    (1, 2), (-1, -1), 'RIGHT')]
    else:
        cmds += [('FONTNAME', (0, 2), (-1, -1), 'Courier'),
                 ('ALIGN',    (0, 2), (-1, -1), 'RIGHT')]

    for i in range(n_rows):
        if i % 2 == 1:
            cmds.append(('BACKGROUND', (0, i + 2), (-1, i + 2), _GRIS_LIGNE))

    if extra:
        cmds.extend(extra)

    t = Table(data, colWidths=col_widths, style=TableStyle(cmds), repeatRows=2)
    out = [t]
    if note:
        out.append(_nt(f'Note : {note}'))
    return out


# ── Correlogramme figure + table numerique (30 lags) ─────────────────────────

def _correl(serie: pd.Series, lags: int = 30, titre: str = '') -> list:
    """Figure ACF/PACF + tableau numerique Lag|AC|PAC|Q-Stat|Prob."""
    from statsmodels.tsa.stattools import acf, pacf
    from statsmodels.stats.diagnostic import acorr_ljungbox

    s = serie.dropna()
    n = len(s)
    if n < lags + 5:
        return [_nt(f'Serie trop courte pour correlogramme ({n} obs).')]

    conf = 1.96 / np.sqrt(n)
    ac   = acf(s, nlags=lags, fft=True)[1:]
    pac  = pacf(s, nlags=lags)[1:]
    lb   = acorr_ljungbox(s, lags=list(range(1, lags + 1)), return_df=True)
    qs   = lb['lb_stat'].values
    pvs  = lb['lb_pvalue'].values

    # Figure
    c_ac  = ['steelblue'  if abs(v) < conf else '#C00000' for v in ac]
    c_pac = ['darkorange' if abs(v) < conf else '#C00000' for v in pac]
    lags_arr = np.arange(1, lags + 1)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 3.5))
    if titre:
        fig.suptitle(titre, fontsize=9, fontweight='bold')
    for ax, vals, cols, ylabel, ttl in [
        (ax1, ac,  c_ac,  'AC',  'Fonction d\'autocorrelation (ACF)'),
        (ax2, pac, c_pac, 'PAC', 'Autocorrelation partielle (PACF)'),
    ]:
        ax.bar(lags_arr, vals, color=cols, width=0.7, alpha=0.85)
        ax.axhline( conf, ls='--', lw=0.8, color='gray', alpha=0.5)
        ax.axhline(-conf, ls='--', lw=0.8, color='gray', alpha=0.5)
        ax.axhline(0, color='black', lw=0.4)
        ax.set_title(ttl, fontsize=9)
        ax.set_xlabel('Retard', fontsize=8)
        ax.set_ylabel(ylabel, fontsize=8)
        ax.tick_params(labelsize=7)
        ax.grid(True, alpha=0.2, linestyle=':')
    fig.tight_layout()
    img = _embed(fig, 15.5, 4.0)

    # Table numerique
    extra = []
    rows  = []
    for k in range(lags):
        pv = pvs[k]
        rows.append([str(k + 1), f'{ac[k]:.3f}', f'{pac[k]:.3f}',
                     f'{qs[k]:.4f}', _fmt_pval(pv)])
        if pv < 0.05:
            ri = k + 2
            extra += [('FONTNAME',  (4, ri), (4, ri), 'Courier-Bold'),
                      ('TEXTCOLOR', (4, ri), (4, ri), _ROUGE_SIG)]

    cw = [LARG_U * 0.10, LARG_U * 0.18, LARG_U * 0.18,
          LARG_U * 0.28, LARG_U * 0.26]
    tabs = _tab('Correlogramme -- ' + (titre or 'serie'),
                ['Retard', 'AC', 'PAC', 'Q-Stat', 'Prob.'],
                rows, col_widths=cw, extra=extra, col0_label=False,
                note=f'Bandes de confiance +/-1.96/sqrt({n}) = {conf:.3f}. '
                     f'Prob. en rouge = significative au seuil 5%.')
    return [img, _sp(0.2)] + tabs


# ── Re-estimation ARIMA pour table de coefficients ────────────────────────────

def _estimer_arima(rendements: pd.Series, p: int, d: int, q: int):
    """
    Re-estime ARIMA(p,d,q) pour extraire la table de coefficients complete.

    Utilise statsmodels ARIMA (meme implementation que arima_selector).
    Retourne le resultat fitté, ou None si echec.
    """
    from statsmodels.tsa.arima.model import ARIMA
    try:
        return ARIMA(rendements.dropna(), order=(p, d, q)).fit()
    except Exception as e:
        _log.warning('  [export_acad] ARIMA re-fit ECHEC : %s', e)
        return None


# ── Section 1 : Statistiques descriptives ────────────────────────────────────

def _sec_stats_desc(log_prix: pd.Series, rendements: pd.Series,
                    ticker: str, freq: str) -> list:
    freq_label = {'daily': 'journaliers', 'weekly': 'hebdomadaires',
                  'monthly': 'mensuels'}.get(freq, freq)
    story = [_h(1, '1. Statistiques descriptives')]

    for serie, label in [(log_prix, f'Log-prix ({ticker})'),
                         (rendements, f'Rendements {freq_label} ({ticker})')]:
        sd = stats_desc(serie)
        jb_str = _fmt(sd.get('jb_stat'), 4) + '  ' + _fmt_signif(sd.get('jb_pval', 1.0))
        rows = [
            ['Observations',    str(int(sd.get('n', 0)))],
            ['Moyenne',         _fmt(sd.get('mean'),     6)],
            ['Mediane',         _fmt(sd.get('median'),   6)],
            ['Maximum',         _fmt(sd.get('max'),      6)],
            ['Minimum',         _fmt(sd.get('min'),      6)],
            ['Ecart-type',      _fmt(sd.get('std'),      6)],
            ['Skewness',        _fmt(sd.get('skew'),     6)],
            ['Kurtosis (exc.)', _fmt(sd.get('kurt_exc'), 6)],
            ['Jarque-Bera',     jb_str],
            ['Prob. JB',        _fmt_pval(sd.get('jb_pval'))],
        ]
        extra = []
        jb_p = sd.get('jb_pval', 1.0)
        if isinstance(jb_p, float) and jb_p < 0.05:
            extra = [('FONTNAME',  (1, 11), (1, 11), 'Courier-Bold'),
                     ('TEXTCOLOR', (1, 11), (1, 11), _ROUGE_SIG)]
        story += _tab(f'Statistiques descriptives -- {label}',
                      ['Statistique', 'Valeur'],
                      rows, extra=extra,
                      col_widths=[LARG_U * 0.60, LARG_U * 0.40])
        story.append(_sp(0.5))
    return story


# ── Section 2 : Correlogrammes ────────────────────────────────────────────────

def _sec_correlogrammes(log_prix: pd.Series, rendements: pd.Series,
                        ticker: str, lags: int = 30) -> list:
    story = [_pb(), _h(1, '2. Correlogrammes')]
    story.append(_h(2, f'2.1 Log-prix ({ticker})'))
    story += _correl(log_prix, lags, f'Log-prix {ticker}')
    story.append(_sp(0.5))
    story.append(_h(2, f'2.2 Rendements ({ticker})'))
    story += _correl(rendements, lags, f'Rendements {ticker}')
    return story


# ── Section 3 : Tests de racine unitaire ─────────────────────────────────────

def _sec_tests_stationnarite(log_prix: pd.Series,
                             rendements: pd.Series, ticker: str) -> list:
    story = [_pb(), _h(1, '3. Tests de racine unitaire')]

    adf_lp  = adf_complet(log_prix)
    pp_lp   = pp_complet(log_prix)
    kpss_lp = kpss_complet(log_prix)
    adf_rd  = adf_complet(rendements)
    pp_rd   = pp_complet(rendements)
    kpss_rd = kpss_complet(rendements)

    story.append(_h(2, f'3.1 Test ADF (Augmented Dickey-Fuller)'))
    story += _tab_adf(adf_lp, f'Log-prix {ticker}',
                      note_nc='H0 : racine unitaire (non-stationnaire).')
    story.append(_sp(0.3))
    story += _tab_adf(adf_rd, f'Rendements {ticker}',
                      note_nc='H0 : racine unitaire (non-stationnaire).')

    story.append(_sp(0.5))
    story.append(_h(2, '3.2 Test PP (Phillips-Perron, 1988)'))
    story += _tab_pp(pp_lp, f'Log-prix {ticker}')
    story.append(_sp(0.3))
    story += _tab_pp(pp_rd, f'Rendements {ticker}')

    story.append(_sp(0.5))
    story.append(_h(2, '3.3 Test KPSS (Kwiatkowski et al., 1992)'))
    story += _tab_kpss(kpss_lp, f'Log-prix {ticker}')
    story.append(_sp(0.3))
    story += _tab_kpss(kpss_rd, f'Rendements {ticker}')

    story.append(_sp(0.5))
    story.append(_h(2, '3.4 Synthese'))
    story += _tab_synthese(adf_lp, pp_lp, kpss_lp, adf_rd, pp_rd, kpss_rd)
    return story


# ── Section 4 : Modélisation ARIMA ───────────────────────────────────────────

def _grille_arima_pivot(df_arima: pd.DataFrame,
                        p_max: int, q_max: int) -> list:
    """
    Tableau p x q des valeurs AIC (format EViews TD).

    Colonnes = AR(p), lignes = MA(q). La cellule du meilleur modele
    est marquee en gras rouge.
    """
    best_aic = df_arima['AIC'].min()

    cols = ['MA \\ AR'] + [str(p) for p in range(0, p_max + 1)]
    rows = []
    extra = []

    for q in range(0, q_max + 1):
        row = [f'MA({q})']
        for p in range(0, p_max + 1):
            if p == 0 and q == 0:
                row.append('—')
            else:
                sub = df_arima[(df_arima['p'] == p) & (df_arima['q'] == q)]
                if sub.empty:
                    row.append('n/a')
                else:
                    val = sub.iloc[0]['AIC']
                    row.append(f'{val:.4f}')
                    if abs(val - best_aic) < 1e-9:
                        ri = q + 2
                        ci = p + 1
                        extra += [
                            ('FONTNAME',  (ci, ri), (ci, ri), 'Courier-Bold'),
                            ('TEXTCOLOR', (ci, ri), (ci, ri), _ROUGE_SIG),
                            ('BACKGROUND',(ci, ri), (ci, ri),
                             colors.HexColor('#FFF2CC')),
                        ]
        rows.append(row)

    n_cols = p_max + 2
    col_widths = [LARG_U * 0.14] + [LARG_U * 0.86 / (p_max + 1)] * (p_max + 1)

    return _tab(
        'Grille de selection ARIMA -- Critere AIC (normalise / observations)',
        cols, rows, col_widths=col_widths,
        extra=extra, col0_label=True,
        note='Valeur en rouge = minimum AIC (modele retenu). '
             'Selection : tous_sig + parcimonie, Burnham-Anderson (2002).',
    )


def _tab_coeff_arima(fit, p: int, d: int, q: int, n: int) -> list:
    """
    Table EViews des coefficients ARIMA estimés.
    Variable | Coefficient | Std. Error | t-Statistique | Prob.
    """
    params  = fit.params
    bse     = fit.bse
    tvals   = fit.tvalues
    pvals   = fit.pvalues

    rows  = []
    extra = []
    ri    = 2  # ligne tableau (titre=0, header=1, data=2...)

    for name in params.index:
        label = name
        if name == 'const':
            label = 'Constante'
        elif name.startswith('ar.L'):
            lag = name.replace('ar.L', '')
            label = f'AR({lag})'
        elif name.startswith('ma.L'):
            lag = name.replace('ma.L', '')
            label = f'MA({lag})'
        elif name == 'sigma2':
            label = 'sigma2'

        pv = pvals[name]
        pv_str = _fmt_pval(pv) + '  ' + _fmt_signif(pv)
        rows.append([label, _fmt(params[name], 6), _fmt(bse[name], 6),
                     _fmt(tvals[name], 4), pv_str])
        if pv < 0.05:
            extra += [('FONTNAME',  (4, ri), (4, ri), 'Courier-Bold'),
                      ('TEXTCOLOR', (4, ri), (4, ri), _ROUGE_SIG)]
        ri += 1

    # Statistiques de regression
    rows.append(['', '', '', '', ''])
    rows.append(['Log-likelihood', _fmt(fit.llf, 4), '',
                 'AIC (normalise)', _fmt(fit.aic / n if n else fit.aic, 6)])
    rows.append(['Observations', str(n), '',
                 'BIC (normalise)', _fmt(fit.bic / n if n else fit.bic, 6)])

    cw = [LARG_U * 0.22, LARG_U * 0.18, LARG_U * 0.18,
          LARG_U * 0.22, LARG_U * 0.20]
    return _tab(
        f'Estimation ARIMA({p},{d},{q}) -- Variable dependante : rendements',
        ['Variable', 'Coefficient', 'Ecart-type', 't-Statistique', 'Prob.'],
        rows, col_widths=cw, extra=extra,
        note='*** p<1%, ** p<5%, * p<10%. Methode : MLE (statsmodels).',
    )


def _sec_arima(rendements: pd.Series, arima_result: dict,
               lags_correl: int = 30) -> list:
    p, d, q = arima_result['p_opt'], arima_result['d_opt'], arima_result['q_opt']
    df_arima = arima_result['df']
    p_max = int(df_arima['p'].max())
    q_max = int(df_arima['q'].max())
    motif  = arima_result.get('motif_selection', '')

    story = [_pb(), _h(1, '4. Modelisation ARIMA (Box-Jenkins)')]
    story.append(_h(2, '4.1 Grille de selection AIC'))
    story += _grille_arima_pivot(df_arima, p_max, q_max)
    story.append(_sp(0.5))

    story.append(_h(2, f'4.2 Estimation du modele retenu : ARIMA({p},{d},{q})'))
    story.append(Paragraph(
        f'Modele selectionne : ARIMA({p},{d},{q}) | Motif : {motif} | '
        f'AIC = {arima_result["aic"]:.6f}',
        _ST['Body']))
    story.append(_sp(0.3))

    fit = _estimer_arima(rendements, p, d, q)
    if fit is not None:
        n_obs = int(fit.nobs)
        story += _tab_coeff_arima(fit, p, d, q, n_obs)
        resid = fit.resid.dropna()
        story.append(_sp(0.5))
        story.append(_h(2, '4.3 Correlogramme des residus ARIMA'))
        story += _correl(pd.Series(resid), lags_correl,
                         f'Residus ARIMA({p},{d},{q})')
        story.append(_sp(0.5))
        story.append(_h(2, '4.4 Correlogramme des carres des residus'))
        story += _correl(pd.Series(resid ** 2), lags_correl,
                         f'Residus^2 ARIMA({p},{d},{q})')
        return story, resid
    else:
        story.append(_nt('Re-estimation ARIMA impossible — table indisponible.'))
        return story, np.array([])


# ── Section 5 : Test ARCH-LM ─────────────────────────────────────────────────

def _sec_arch_lm(resid: np.ndarray, lags_list: list) -> list:
    story = [_pb(), _h(1, '5. Test ARCH-LM sur residus ARIMA (Engle, 1982)')]

    resultats = arch_lm(resid, lags=lags_list)
    rows, extra = [], []
    ri = 2
    for r in resultats:
        lag    = r['lag']
        fp_str = _fmt_pval(r.get('f_pval'))
        lp_str = _fmt_pval(r.get('lm_pval'))
        verdict = 'ARCH detecte ***' if r.get('lm_pval', 1.0) < 0.05 else 'Pas d\'ARCH'
        rows.append([str(lag),
                     _fmt(r.get('f_stat'),  4),
                     fp_str,
                     _fmt(r.get('lm_stat'), 4),
                     lp_str,
                     verdict])
        if r.get('lm_pval', 1.0) < 0.05:
            extra += [('FONTNAME',  (5, ri), (5, ri), 'Courier-Bold'),
                      ('TEXTCOLOR', (5, ri), (5, ri), _ROUGE_SIG)]
        ri += 1

    cw = [LARG_U * 0.08, LARG_U * 0.14, LARG_U * 0.14,
          LARG_U * 0.14, LARG_U * 0.14, LARG_U * 0.36]
    story += _tab(
        'Test d\'heteroscedasticite ARCH-LM (Engle 1982)',
        ['Lags', 'F-stat', 'Prob. F', 'LM (obs*R2)', 'Prob. LM', 'Verdict'],
        rows, col_widths=cw, extra=extra, col0_label=False,
        note='H0 : absence d\'effet ARCH (homosc.). Rejet si p < 5%.',
    )
    return story


# ── Section 6 : Modélisation GARCH ───────────────────────────────────────────

def _grille_garch(df_garch: pd.DataFrame, top_n: int = 15) -> list:
    """Table comparative des modeles GARCH (top_n par AIC)."""
    df_top = df_garch.nsmallest(top_n, 'AIC_norm').copy()

    rows, extra = [], []
    ri = 2
    best_aic = float(df_top['AIC_norm'].iloc[0])

    for _, r in df_top.iterrows():
        modele = str(r['modele'])
        spec   = f'({int(r["p"])},{int(r["o"])},{int(r["q"])})'
        dist   = str(r.get('dist', ''))
        aic    = _fmt(r.get('AIC_norm'),  6)
        bic    = _fmt(r.get('BIC_norm'),  6)
        pers   = _fmt(r.get('persistance'), 4)
        sig    = 'Oui' if r.get('tous_sig_vol', False) else 'Non'

        rows.append([f'{modele}{spec}', dist, aic, bic, pers, sig])

        if abs(float(r.get('AIC_norm', 1e9)) - best_aic) < 1e-9:
            for c in range(6):
                extra += [('FONTNAME',  (c, ri), (c, ri), 'Courier-Bold'),
                          ('TEXTCOLOR', (c, ri), (c, ri), _ROUGE_SIG)]
        ri += 1

    cw = [LARG_U * 0.28, LARG_U * 0.10, LARG_U * 0.15,
          LARG_U * 0.15, LARG_U * 0.16, LARG_U * 0.16]
    return _tab(
        f'Grille comparative GARCH -- Top {top_n} modeles par AIC',
        ['Specification', 'Dist.', 'AIC/n', 'BIC/n', 'Persistance', 'Vol. sig.'],
        rows, col_widths=cw, extra=extra,
        note='Modele en rouge = meilleur AIC. Vol. sig. = tous parametres volatilite p<5%.',
    )


def _tab_coeff_garch(garch_final, best: pd.Series) -> list:
    """
    Table EViews de l'equation de variance GARCH estimée.
    Variable | Coefficient | Ecart-type | t-Stat | Prob.
    """
    params = garch_final.params
    bse    = garch_final.std_err
    pvals  = garch_final.pvalues
    tvals  = params / bse.replace(0, np.nan)

    nom_modele = str(best.get('modele', ''))
    dist       = str(best.get('dist',   ''))
    p_, o_, q_ = int(best.get('p', 1)), int(best.get('o', 0)), int(best.get('q', 1))

    # Libelles lisibles
    label_map = {
        'mu':    'mu (moyenne)',
        'omega': 'omega (constante variance)',
        'alpha[1]': 'alpha[1] (ARCH, eps^2_{t-1})',
        'alpha[2]': 'alpha[2] (ARCH, eps^2_{t-2})',
        'gamma[1]': 'gamma[1] (asymetrie, effet levier)',
        'gamma[2]': 'gamma[2] (asymetrie)',
        'beta[1]':  'beta[1] (GARCH, sigma^2_{t-1})',
        'beta[2]':  'beta[2] (GARCH, sigma^2_{t-2})',
        'beta[3]':  'beta[3] (GARCH, sigma^2_{t-3})',
        'eta':      'eta (skewness)',
        'lambda':   'lambda (shape / degres liberte)',
        'nu':       'nu (degres liberte)',
    }

    rows, extra = [], []
    ri = 2
    for name in params.index:
        label = label_map.get(name, name)
        pv    = float(pvals.get(name, 1.0))
        tv    = float(tvals.get(name, np.nan))
        pv_str = _fmt_pval(pv) + '  ' + _fmt_signif(pv)
        rows.append([label, _fmt(float(params[name]), 6),
                     _fmt(float(bse[name]), 6),
                     _fmt(tv, 4), pv_str])
        if pv < 0.05:
            extra += [('FONTNAME',  (4, ri), (4, ri), 'Courier-Bold'),
                      ('TEXTCOLOR', (4, ri), (4, ri), _ROUGE_SIG)]
        ri += 1

    # Infos de convergence
    rows.append(['', '', '', '', ''])
    rows.append(['Log-likelihood', _fmt(garch_final.loglikelihood, 4), '',
                 'AIC (normalise)', _fmt(best.get('AIC_norm', np.nan), 6)])
    rows.append(['Persistance', _fmt(best.get('persistance', np.nan), 4), '',
                 'BIC (normalise)', _fmt(best.get('BIC_norm', np.nan), 6)])

    cw = [LARG_U * 0.35, LARG_U * 0.15, LARG_U * 0.15,
          LARG_U * 0.15, LARG_U * 0.20]
    titre = (f'Equation de variance -- {nom_modele}({p_},{o_},{q_}) '
             f'[{dist}]')
    return _tab(titre,
                ['Variable', 'Coefficient', 'Ecart-type', 't-Stat', 'Prob.'],
                rows, col_widths=cw, extra=extra,
                note='*** p<1%, ** p<5%, * p<10%. Methode : MV (arch library).')


def _tab_stats_std_resid(std_resid: np.ndarray) -> list:
    """Statistiques descriptives des residus standardises z_t = eps_t/sigma_t."""
    sd = stats_desc(pd.Series(std_resid))
    jb_str = _fmt(sd.get('jb_stat'), 4) + '  ' + _fmt_signif(sd.get('jb_pval', 1.0))
    rows = [
        ['Observations',    str(int(sd.get('n', 0)))],
        ['Moyenne',         _fmt(sd.get('mean'),     6)],
        ['Mediane',         _fmt(sd.get('median'),   6)],
        ['Maximum',         _fmt(sd.get('max'),      6)],
        ['Minimum',         _fmt(sd.get('min'),      6)],
        ['Ecart-type',      _fmt(sd.get('std'),      6)],
        ['Skewness',        _fmt(sd.get('skew'),     6)],
        ['Kurtosis (exc.)', _fmt(sd.get('kurt_exc'), 6)],
        ['Jarque-Bera',     jb_str],
        ['Prob. JB',        _fmt_pval(sd.get('jb_pval'))],
    ]
    extra = []
    jbp = sd.get('jb_pval', 1.0)
    if isinstance(jbp, float) and jbp < 0.05:
        extra = [('FONTNAME',  (1, 11), (1, 11), 'Courier-Bold'),
                 ('TEXTCOLOR', (1, 11), (1, 11), _ROUGE_SIG)]
    return _tab('Statistiques descriptives -- Residus standardises z_t',
                ['Statistique', 'Valeur'], rows, extra=extra,
                col_widths=[LARG_U * 0.60, LARG_U * 0.40])


def _sec_garch(df_garch: pd.DataFrame, garch_final, best: pd.Series,
               lags_correl: int = 30, lags_arch: list = None) -> list:
    if lags_arch is None:
        lags_arch = [1, 4, 8, 12]

    p_, o_, q_ = int(best['p']), int(best['o']), int(best['q'])
    dist_ = str(best.get('dist', ''))
    modele_ = str(best.get('modele', ''))

    story = [_pb(), _h(1, '6. Modelisation GARCH')]
    story.append(_h(2, '6.1 Grille comparative des specifications'))
    story += _grille_garch(df_garch)
    story.append(_sp(0.5))

    story.append(_h(2, f'6.2 Estimation : {modele_}({p_},{o_},{q_}) [{dist_}]'))
    story += _tab_coeff_garch(garch_final, best)
    story.append(_sp(0.5))

    std_resid = np.asarray(garch_final.std_resid)
    std_resid = std_resid[~np.isnan(std_resid)]

    story.append(_h(2, '6.3 Statistiques des residus standardises'))
    story += _tab_stats_std_resid(std_resid)
    story.append(_sp(0.5))

    story.append(_h(2, '6.4 Correlogramme des residus standardises z_t'))
    story += _correl(pd.Series(std_resid), lags_correl,
                     f'Residus standardises {modele_}({p_},{o_},{q_})')
    story.append(_sp(0.5))

    story.append(_h(2, '6.5 Correlogramme des carres des residus standardises z^2_t'))
    story += _correl(pd.Series(std_resid ** 2), lags_correl,
                     f'Residus standardises^2 {modele_}({p_},{o_},{q_})')
    story.append(_sp(0.5))

    story.append(_h(2, '6.6 Test ARCH-LM sur residus standardises'))
    resultats = arch_lm(std_resid, lags=lags_arch)
    rows, extra = [], []
    ri = 2
    for r in resultats:
        verdict = 'ARCH residuel ***' if r.get('lm_pval', 1.0) < 0.05 else 'OK'
        rows.append([str(r['lag']), _fmt(r.get('f_stat'), 4),
                     _fmt_pval(r.get('f_pval')),
                     _fmt(r.get('lm_stat'), 4),
                     _fmt_pval(r.get('lm_pval')), verdict])
        if r.get('lm_pval', 1.0) < 0.05:
            extra += [('FONTNAME',  (5, ri), (5, ri), 'Courier-Bold'),
                      ('TEXTCOLOR', (5, ri), (5, ri), _ROUGE_SIG)]
        ri += 1
    cw = [LARG_U * 0.08, LARG_U * 0.14, LARG_U * 0.14,
          LARG_U * 0.14, LARG_U * 0.14, LARG_U * 0.36]
    story += _tab('Test ARCH-LM -- Residus standardises z_t',
                  ['Lags', 'F-stat', 'Prob. F', 'LM (obs*R2)', 'Prob. LM', 'Verdict'],
                  rows, col_widths=cw, extra=extra, col0_label=False,
                  note='H0 : absence d\'effet ARCH residuel dans z_t.')
    return story


# ── Section 7 : VaR & TVaR ───────────────────────────────────────────────────

def _sec_var(df_var: pd.DataFrame) -> list:
    story = [_pb(), _h(1, '7. Value at Risk (VaR) et TVaR')]
    story.append(Paragraph(
        'Niveaux de confiance : 90%, 95%, 99%. Methodes : historique, '
        'normale, Student, Cornish-Fisher, Monte-Carlo, GARCH dynamique.',
        _ST['Body']))
    story.append(_sp(0.3))

    if df_var is None or df_var.empty:
        story.append(_nt('Tableau VaR non disponible.'))
        return story

    df = df_var.reset_index()
    col_val = [c for c in df.columns if c not in ('Niveau', 'index')]
    cols = ['Niveau'] + col_val
    rows = []
    for _, row in df.iterrows():
        niveau = str(row['Niveau']) if 'Niveau' in row else str(row.get('index', ''))
        data_row = [niveau] + [_fmt(row.get(c), 4) for c in col_val]
        rows.append(data_row)

    cw_first = LARG_U * 0.10
    cw_rest  = (LARG_U - cw_first) / len(col_val)
    col_widths = [cw_first] + [cw_rest] * len(col_val)
    story += _tab('Value at Risk et TVaR (%)',
                  cols, rows, col_widths=col_widths, col0_label=True,
                  note='Valeurs en % du portefeuille. Pertes = valeurs negatives.')
    return story


# ── Section 8 : Backtesting ───────────────────────────────────────────────────

def _sec_backtest(df_bt: pd.DataFrame, T_train: int, T_eff_dyn: int) -> list:
    story = [_pb(), _h(1, '8. Backtesting (validation hors-echantillon)')]
    story.append(Paragraph(
        f'Echantillon train : {T_train} observations. '
        f'Echantillon test GARCH dynamique : {T_eff_dyn} observations. '
        'Tests : Kupiec (LR_UC) et Christoffersen (LR_CC).',
        _ST['Body']))
    story.append(_sp(0.3))

    if df_bt is None or df_bt.empty:
        story.append(_nt('Tableau backtesting non disponible.'))
        return story

    df = df_bt.reset_index(drop=True)
    cols_fixed = ['Methode', 'Niveau', 'N viol.', 'Taux obs.',
                  'p_UC', 'p_CC', 'Verdict CC']
    available = [c for c in cols_fixed if c in df.columns]
    rows, extra = [], []
    ri = 2
    for _, row in df.iterrows():
        data_row = []
        for c in available:
            val = row.get(c, '')
            if c in ('p_UC', 'p_CC'):
                data_row.append(_fmt_pval(val))
            elif c in ('Taux obs.',):
                data_row.append(_fmt(val, 4))
            else:
                data_row.append(str(val))
        rows.append(data_row)

        verdict = str(row.get('Verdict CC', ''))
        if verdict.upper() not in ('OK', ''):
            extra += [('FONTNAME',  (len(available) - 1, ri),
                       (len(available) - 1, ri), 'Courier-Bold'),
                      ('TEXTCOLOR', (len(available) - 1, ri),
                       (len(available) - 1, ri), _ROUGE_SIG)]
        ri += 1

    cw_first = LARG_U * 0.22
    cw_rest  = (LARG_U - cw_first) / (len(available) - 1)
    col_widths = [cw_first] + [cw_rest] * (len(available) - 1)
    story += _tab('Backtesting OOS -- Tests Kupiec (1995) et Christoffersen (1998)',
                  available, rows, col_widths=col_widths, extra=extra,
                  note='H0 Kupiec (UC) : taux de violations = (1-alpha). '
                       'H0 Christoffersen (CC) : UC + independance. '
                       'p_UC, p_CC : p-valeurs LR.')
    return story


# ── Document PDF ──────────────────────────────────────────────────────────────

class _AcademicDoc(BaseDocTemplate):
    """PDF avec pied de page (page N / N_total)."""

    def __init__(self, filename: str, titre: str, ticker: str):
        BaseDocTemplate.__init__(
            self, filename,
            pagesize=A4,
            leftMargin=MARGE_G, rightMargin=MARGE_D,
            topMargin=MARGE_H, bottomMargin=MARGE_B,
        )
        self._titre  = titre
        self._ticker = ticker
        frame = Frame(
            MARGE_G, MARGE_B,
            A4[0] - MARGE_G - MARGE_D,
            A4[1] - MARGE_H - MARGE_B,
            id='normal',
        )
        self.addPageTemplates([
            PageTemplate(id='all', frames=frame, onPage=self._footer),
        ])

    def _footer(self, canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor('#595959'))
        canvas.drawString(
            MARGE_G,
            MARGE_B * 0.5,
            f'{self._ticker} — Rapport academique — '
            f'{datetime.today().strftime("%d/%m/%Y")}',
        )
        canvas.drawRightString(
            A4[0] - MARGE_D,
            MARGE_B * 0.5,
            f'Page {doc.page}',
        )
        canvas.restoreState()


def _page_garde(ticker: str, start_date: str, end_date: str,
                freq: str, best: pd.Series) -> list:
    st = _ST
    freq_label = {'daily': 'Journalier', 'weekly': 'Hebdomadaire',
                  'monthly': 'Mensuel'}.get(freq, freq.capitalize())
    p_, o_, q_ = int(best.get('p', 1)), int(best.get('o', 0)), int(best.get('q', 1))
    modele_str = (f"{best.get('modele', 'GARCH')}({p_},{o_},{q_})"
                  f"[{best.get('dist', 'normal')}]")

    story = [_sp(3)]
    story.append(Paragraph('ANALYSE DE LA VOLATILITE', st['Titre']))
    story.append(Paragraph(f'{nom_actif(ticker)} — {ticker}', st['Titre']))
    story.append(_sp(1))
    story.append(Paragraph(
        f'Periode : {start_date} — {end_date} | Frequence : {freq_label}',
        st['SousTitre']))
    story.append(Paragraph(
        f'Modele ARIMA : ({best.get("arima_p","?")},'
        f'{best.get("arima_d","?")},{best.get("arima_q","?")}) | '
        f'Modele GARCH : {modele_str}',
        st['SousTitre']))
    story.append(_sp(1.5))
    story.append(Paragraph(
        f'Genere le {datetime.today().strftime("%d %B %Y")}',
        st['SousTitre']))
    story.append(_pb())
    return story


# ── Excel ─────────────────────────────────────────────────────────────────────

def _excel_stats_desc(ws, log_prix: pd.Series, rendements: pd.Series) -> None:
    """Feuille Excel : statistiques descriptives."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    bleu  = PatternFill('solid', fgColor='1F3864')
    clair = PatternFill('solid', fgColor='BDD7EE')
    gris  = PatternFill('solid', fgColor='F2F2F2')

    def _header_cell(cell, text):
        cell.value = text
        cell.font  = Font(bold=True, color='FFFFFF')
        cell.fill  = bleu
        cell.alignment = Alignment(horizontal='center')

    stats_all = {}
    for serie, label in [(log_prix, 'Log-prix'),
                         (rendements, 'Rendements')]:
        sd = stats_desc(serie)
        stats_all[label] = sd

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    _header_cell(ws.cell(1, 1), 'Statistique')
    _header_cell(ws.cell(1, 2), 'Log-prix')
    _header_cell(ws.cell(1, 3), 'Rendements')

    labels = [('Observations', 'n', lambda x: str(int(x))),
              ('Moyenne',      'mean',     lambda x: f'{x:.6f}'),
              ('Mediane',      'median',   lambda x: f'{x:.6f}'),
              ('Maximum',      'max',      lambda x: f'{x:.6f}'),
              ('Minimum',      'min',      lambda x: f'{x:.6f}'),
              ('Ecart-type',   'std',      lambda x: f'{x:.6f}'),
              ('Skewness',     'skew',     lambda x: f'{x:.6f}'),
              ('Kurtosis exc.','kurt_exc', lambda x: f'{x:.6f}'),
              ('Jarque-Bera',  'jb_stat',  lambda x: f'{x:.4f}'),
              ('Prob. JB',     'jb_pval',  lambda x: f'{x:.4f}'),]

    for row_i, (lbl, key, fmt) in enumerate(labels, start=2):
        cell_l = ws.cell(row_i, 1)
        cell_l.value = lbl
        if row_i % 2 == 1:
            cell_l.fill = gris
        for col_i, name in enumerate(('Log-prix', 'Rendements'), start=2):
            val = stats_all[name].get(key, float('nan'))
            cell = ws.cell(row_i, col_i)
            try:
                cell.value = fmt(float(val))
            except Exception:
                cell.value = 'N/A'
            if row_i % 2 == 1:
                cell.fill = gris
            cell.alignment = Alignment(horizontal='right')


def _excel_correl(ws, serie: pd.Series, lags: int, label: str) -> None:
    """Feuille Excel : correlogramme numerique."""
    from statsmodels.tsa.stattools import acf, pacf
    from statsmodels.stats.diagnostic import acorr_ljungbox
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    bleu  = PatternFill('solid', fgColor='1F3864')
    gris  = PatternFill('solid', fgColor='F2F2F2')

    s   = serie.dropna()
    n   = len(s)
    ac  = acf(s, nlags=lags, fft=True)[1:]
    pac = pacf(s, nlags=lags)[1:]
    lb  = acorr_ljungbox(s, lags=list(range(1, lags + 1)), return_df=True)
    qs  = lb['lb_stat'].values
    pvs = lb['lb_pvalue'].values

    headers = ['Retard', 'AC', 'PAC', 'Q-Stat', 'Prob.']
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(1, col_i)
        cell.value = h
        cell.font  = Font(bold=True, color='FFFFFF')
        cell.fill  = bleu
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64 + col_i)].width = 14

    for k in range(lags):
        row_i = k + 2
        ws.cell(row_i, 1).value = k + 1
        ws.cell(row_i, 2).value = round(float(ac[k]),  4)
        ws.cell(row_i, 3).value = round(float(pac[k]), 4)
        ws.cell(row_i, 4).value = round(float(qs[k]),  4)
        ws.cell(row_i, 5).value = round(float(pvs[k]), 4)
        if k % 2 == 1:
            for c in range(1, 6):
                ws.cell(row_i, c).fill = gris
        if pvs[k] < 0.05:
            from openpyxl.styles import Font as _Font
            ws.cell(row_i, 5).font = _Font(bold=True, color='C00000')


def _excel_arima_grille(ws, df_arima: pd.DataFrame,
                        p_max: int, q_max: int) -> None:
    """Feuille Excel : grille AIC p*q."""
    from openpyxl.styles import Font, Alignment, PatternFill

    bleu   = PatternFill('solid', fgColor='1F3864')
    jaune  = PatternFill('solid', fgColor='FFF2CC')
    best_a = df_arima['AIC'].min()

    ws.cell(1, 1).value = 'MA \\ AR'
    ws.cell(1, 1).font  = Font(bold=True, color='FFFFFF')
    ws.cell(1, 1).fill  = bleu

    for p in range(0, p_max + 1):
        cell = ws.cell(1, p + 2)
        cell.value = f'AR({p})'
        cell.font  = Font(bold=True, color='FFFFFF')
        cell.fill  = bleu
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(65 + p + 1)].width = 12

    ws.column_dimensions['A'].width = 10
    for q in range(0, q_max + 1):
        ri = q + 2
        ws.cell(ri, 1).value = f'MA({q})'
        ws.cell(ri, 1).font  = Font(bold=True)
        ws.cell(ri, 1).fill  = bleu
        ws.cell(ri, 1).font  = Font(bold=True, color='FFFFFF')
        for p in range(0, p_max + 1):
            ci = p + 2
            if p == 0 and q == 0:
                ws.cell(ri, ci).value = '—'
            else:
                sub = df_arima[(df_arima['p'] == p) & (df_arima['q'] == q)]
                if sub.empty:
                    ws.cell(ri, ci).value = 'n/a'
                else:
                    val = float(sub.iloc[0]['AIC'])
                    ws.cell(ri, ci).value = round(val, 4)
                    ws.cell(ri, ci).alignment = Alignment(horizontal='right')
                    if abs(val - best_a) < 1e-9:
                        ws.cell(ri, ci).fill = jaune
                        ws.cell(ri, ci).font = Font(bold=True, color='C00000')


def _excel_garch_grille(ws, df_garch: pd.DataFrame, top_n: int = 20) -> None:
    """Feuille Excel : grille comparative GARCH."""
    from openpyxl.styles import Font, Alignment, PatternFill

    bleu  = PatternFill('solid', fgColor='1F3864')
    gris  = PatternFill('solid', fgColor='F2F2F2')
    jaune = PatternFill('solid', fgColor='FFF2CC')

    df_top    = df_garch.nsmallest(top_n, 'AIC_norm').copy()
    best_aic  = float(df_top['AIC_norm'].iloc[0])
    cols_show = ['modele', 'p', 'o', 'q', 'dist',
                 'AIC_norm', 'BIC_norm', 'persistance', 'tous_sig_vol']
    headers   = ['Modele', 'p', 'o', 'q', 'Dist',
                 'AIC/n', 'BIC/n', 'Persistance', 'Vol. sig.']

    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(1, col_i)
        cell.value = h
        cell.font  = Font(bold=True, color='FFFFFF')
        cell.fill  = bleu
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64 + col_i)].width = 14

    for row_i, (_, r) in enumerate(df_top.iterrows(), start=2):
        is_best = abs(float(r.get('AIC_norm', 1e9)) - best_aic) < 1e-9
        for col_i, key in enumerate(cols_show, 1):
            val  = r.get(key, '')
            cell = ws.cell(row_i, col_i)
            if isinstance(val, (float, np.floating)):
                cell.value = round(float(val), 6)
            elif isinstance(val, (bool, np.bool_)):
                cell.value = 'Oui' if val else 'Non'
            else:
                cell.value = str(val)
            cell.alignment = Alignment(horizontal='center')
            if is_best:
                cell.fill = jaune
                cell.font = Font(bold=True, color='C00000')
            elif row_i % 2 == 1:
                cell.fill = gris


# ── Generateur principal ──────────────────────────────────────────────────────

def generer_export_academique(
    prix,
    rendements: pd.Series,
    arima_result: dict,
    garch_final,
    df_garch: pd.DataFrame,
    df_var,
    df_bt,
    best,
    config: dict,
    T_train: int = 0,
    T_eff_dyn: int = 0,
) -> dict:
    """
    Genere PDF et Excel academiques (tables style EViews) dans exports/.

    Le dossier de sortie (config.export_academique.dossier) n'est JAMAIS
    efface par force_clean. Les fichiers sont horodates pour eviter
    d'ecraser les productions precedentes.

    Parameters
    ----------
    prix, rendements, arima_result, garch_final, df_garch, df_var, df_bt,
    best, config : sorties du pipeline TickerLab (cache ou run direct).
    T_train, T_eff_dyn : tailles echantillons train/test.

    Returns
    -------
    dict
        Cles 'pdf' et 'xlsx' — chemins absolus des fichiers generes.
    """
    ea_cfg  = config.get('export_academique', {})
    dossier = ea_cfg.get('dossier', 'exports')
    lags    = int(ea_cfg.get('lags_correlogramme', 30))
    lags_arch = [int(x) for x in ea_cfg.get('lags_arch_test', [1, 4, 8, 12])]

    out_dir = Path(config['output']['dossier_resultats']).parent / dossier
    out_dir.mkdir(parents=True, exist_ok=True)

    ticker     = config['data']['ticker']
    start_date = config['data']['start_date']
    end_date   = config['data']['end_date']
    freq       = config['data'].get('frequency', 'daily')
    date_str   = datetime.today().strftime('%Y%m%d')

    # Sécuriser nom de fichier
    safe_ticker = ticker.replace('=', '_').replace('^', '_')
    pdf_path  = out_dir / f'Rapport_Academique_{safe_ticker}_{date_str}.pdf'
    xlsx_path = out_dir / f'Tableaux_{safe_ticker}_{date_str}.xlsx'

    import numpy as np
    log_prix = np.log(prix['prix'].dropna())
    log_prix = pd.Series(log_prix.values, index=prix['prix'].dropna().index,
                         name='log_prix')

    best_s = best if hasattr(best, 'get') else pd.Series(best)

    # ── PDF ───────────────────────────────────────────────────────────────────
    _log.info('  [Export acad.] Generation PDF -> %s', pdf_path.name)
    story = []

    # Page de garde
    story += _page_garde(ticker, start_date, end_date, freq, best_s)

    # Section 1 — Stats desc
    story += _sec_stats_desc(log_prix, rendements, ticker, freq)

    # Section 2 — Correlogrammes
    story += _sec_correlogrammes(log_prix, rendements, ticker, lags)

    # Section 3 — Tests stationnarite
    story += _sec_tests_stationnarite(log_prix, rendements, ticker)

    # Section 4 — ARIMA (retourne aussi les residus)
    sec4_result = _sec_arima(rendements, arima_result, lags)
    if isinstance(sec4_result, tuple):
        sec4_story, arima_resid = sec4_result
    else:
        sec4_story, arima_resid = sec4_result, np.array([])
    story += sec4_story

    # Section 5 — Test ARCH sur residus ARIMA
    if len(arima_resid) > 20:
        story += _sec_arch_lm(arima_resid, lags_arch)

    # Section 6 — GARCH
    story += _sec_garch(df_garch, garch_final, best_s, lags, lags_arch)

    # Section 7 — VaR
    story += _sec_var(df_var)

    # Section 8 — Backtesting
    story += _sec_backtest(df_bt, T_train, T_eff_dyn)

    # Compilation PDF
    doc = _AcademicDoc(str(pdf_path), titre='Rapport Academique', ticker=ticker)
    doc.build(story)
    _log.info('  [Export acad.] PDF OK -> %s', pdf_path)

    # ── Excel ─────────────────────────────────────────────────────────────────
    _log.info('  [Export acad.] Generation Excel -> %s', xlsx_path.name)
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # supprimer feuille par defaut

        # Feuille 1 : Stats desc
        ws1 = wb.create_sheet('Stats_Desc')
        _excel_stats_desc(ws1, log_prix, rendements)

        # Feuilles 2-3 : Correlogrammes
        ws2a = wb.create_sheet('Correl_LogPrix')
        _excel_correl(ws2a, log_prix, lags, f'Log-prix {ticker}')
        ws2b = wb.create_sheet('Correl_Rendements')
        _excel_correl(ws2b, rendements, lags, f'Rendements {ticker}')

        # Feuille 4 : Grille ARIMA
        ws3 = wb.create_sheet('ARIMA_Grille')
        df_ar = arima_result['df']
        p_max = int(df_ar['p'].max())
        q_max = int(df_ar['q'].max())
        _excel_arima_grille(ws3, df_ar, p_max, q_max)

        # Feuille 5 : Coefficients ARIMA
        p, d, q = (arima_result['p_opt'], arima_result['d_opt'],
                   arima_result['q_opt'])
        fit = _estimer_arima(rendements, p, d, q)
        if fit is not None:
            ws4 = wb.create_sheet('ARIMA_Coefficients')
            from openpyxl.styles import Font, Alignment, PatternFill
            bleu = PatternFill('solid', fgColor='1F3864')
            headers = ['Variable', 'Coefficient', 'Ecart-type',
                       't-Statistique', 'Prob.']
            for col_i, h in enumerate(headers, 1):
                cell = ws4.cell(1, col_i)
                cell.value = h
                cell.font  = Font(bold=True, color='FFFFFF')
                cell.fill  = bleu
                cell.alignment = Alignment(horizontal='center')
                ws4.column_dimensions[chr(64 + col_i)].width = 18
            for row_i, name in enumerate(fit.params.index, start=2):
                pv  = float(fit.pvalues[name])
                tv  = float(fit.tvalues[name])
                ws4.cell(row_i, 1).value = name
                ws4.cell(row_i, 2).value = round(float(fit.params[name]), 6)
                ws4.cell(row_i, 3).value = round(float(fit.bse[name]), 6)
                ws4.cell(row_i, 4).value = round(tv, 4)
                ws4.cell(row_i, 5).value = round(pv, 4)

        # Feuille 6 : Grille GARCH
        ws5 = wb.create_sheet('GARCH_Grille')
        _excel_garch_grille(ws5, df_garch)

        # Feuille 7 : Coefficients GARCH
        ws6 = wb.create_sheet('GARCH_Coefficients')
        from openpyxl.styles import Font, Alignment, PatternFill
        bleu = PatternFill('solid', fgColor='1F3864')
        gris = PatternFill('solid', fgColor='F2F2F2')
        headers = ['Parametre', 'Coefficient', 'Ecart-type', 't-Stat', 'Prob.']
        for col_i, h in enumerate(headers, 1):
            cell = ws6.cell(1, col_i)
            cell.value = h
            cell.font  = Font(bold=True, color='FFFFFF')
            cell.fill  = bleu
            cell.alignment = Alignment(horizontal='center')
            ws6.column_dimensions[chr(64 + col_i)].width = 20
        params_g = garch_final.params
        bse_g    = garch_final.std_err
        pvals_g  = garch_final.pvalues
        tvals_g  = params_g / bse_g.replace(0, np.nan)
        for row_i, name in enumerate(params_g.index, start=2):
            ws6.cell(row_i, 1).value = name
            ws6.cell(row_i, 2).value = round(float(params_g[name]), 6)
            ws6.cell(row_i, 3).value = round(float(bse_g[name]),    6)
            ws6.cell(row_i, 4).value = round(float(tvals_g.get(name, np.nan)), 4)
            ws6.cell(row_i, 5).value = round(float(pvals_g[name]),  4)
            if row_i % 2 == 1:
                for c in range(1, 6):
                    ws6.cell(row_i, c).fill = gris

        # Feuille 8 : VaR
        if df_var is not None and not df_var.empty:
            ws7 = wb.create_sheet('VaR_TVaR')
            df_var_reset = df_var.reset_index()
            for col_i, col in enumerate(df_var_reset.columns, 1):
                ws7.cell(1, col_i).value = str(col)
                ws7.cell(1, col_i).font  = Font(bold=True, color='FFFFFF')
                ws7.cell(1, col_i).fill  = bleu
                ws7.column_dimensions[chr(64 + col_i)].width = 18
            for row_i, (_, row) in enumerate(df_var_reset.iterrows(), start=2):
                for col_i, val in enumerate(row, 1):
                    try:
                        ws7.cell(row_i, col_i).value = round(float(val), 4)
                    except Exception:
                        ws7.cell(row_i, col_i).value = str(val)

        # Feuille 9 : Backtesting
        if df_bt is not None and not df_bt.empty:
            ws8 = wb.create_sheet('Backtesting')
            df_bt_r = df_bt.reset_index(drop=True)
            for col_i, col in enumerate(df_bt_r.columns, 1):
                ws8.cell(1, col_i).value = str(col)
                ws8.cell(1, col_i).font  = Font(bold=True, color='FFFFFF')
                ws8.cell(1, col_i).fill  = bleu
                ws8.column_dimensions[chr(64 + col_i)].width = 16
            for row_i, (_, row) in enumerate(df_bt_r.iterrows(), start=2):
                for col_i, val in enumerate(row, 1):
                    try:
                        ws8.cell(row_i, col_i).value = round(float(val), 4)
                    except Exception:
                        ws8.cell(row_i, col_i).value = str(val)

        wb.save(str(xlsx_path))
        _log.info('  [Export acad.] Excel OK -> %s', xlsx_path)

    except ImportError:
        _log.warning('  [Export acad.] openpyxl absent — Excel non genere.')

    return {'pdf': str(pdf_path), 'xlsx': str(xlsx_path)}
